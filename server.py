# server.py

from flask import Flask, request, jsonify, send_from_directory, session, Response
from flask_sock import Sock
import json
import os
import io
import threading
import time
import argparse
import socket
import subprocess
import struct
import select
from datetime import datetime
from datetime import timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
from werkzeug.security import generate_password_hash, check_password_hash
from device_manager import DeviceManager
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))
DATA_ROOT = os.path.join(BASE, "data")
USERS_FILE = os.path.join(BASE, "users.json")
PERSONS_FILE = os.path.join(BASE, "persons.json")
DEVICE_MAP_FILE = os.path.join(BASE, "device_map.json")

app = Flask(__name__)
sock = Sock(app)
app.secret_key = os.environ.get('DAHUA_DOOR_WEB_SECRET_KEY') or os.urandom(32)
app.config['SESSION_PERMANENT'] = True
from datetime import timedelta
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)
manager = DeviceManager()

# ================= 工具函数 =================
def ok(data=None):
    return jsonify({"code": 0, "msg": "ok", "data": data})

def err(msg, code=400):
    return jsonify({"code": -1, "msg": str(msg)}), code

def load_users():
    if not os.path.exists(USERS_FILE):
        save_users({})
        return {}
    with open(USERS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_users(users):
    with open(USERS_FILE, "w", encoding="utf-8") as f:
        json.dump(users, f, indent=2, ensure_ascii=False)

def get_user_role(username):
    """返回用户角色：'admin' 或 'user'"""
    users = load_users()
    entry = users.get(username)
    if entry is None:
        return None
    if isinstance(entry, dict):
        return entry.get("role", "user")
    # 旧格式（纯密码哈希字符串），视为普通用户
    return "user"

def get_user_password_hash(username):
    users = load_users()
    entry = users.get(username)
    if entry is None:
        return None
    if isinstance(entry, dict):
        return entry.get("password")
    return entry  # 旧格式

def is_admin(username):
    return get_user_role(username) == "admin"

def require_admin():
    username = session.get("username")
    if not username or not is_admin(username):
        raise Exception("需要管理员权限")

def get_user_dir(username):
    user_dir = os.path.join(DATA_ROOT, username)
    if not os.path.exists(user_dir):
        os.makedirs(user_dir, exist_ok=True)
    return user_dir

# ---------- 全局设备ID映射 ----------
def load_device_map():
    if not os.path.exists(DEVICE_MAP_FILE):
        save_device_map({})
        return {}
    try:
        with open(DEVICE_MAP_FILE, "r", encoding="utf-8") as f:
            content = f.read()
            if not content.strip():
                save_device_map({})
                return {}
            return json.loads(content)
    except (json.JSONDecodeError, ValueError):
        print("device_map.json 格式错误，已重置为空")
        save_device_map({})
        return {}

def save_device_map(mapping):
    with open(DEVICE_MAP_FILE, "w", encoding="utf-8") as f:
        json.dump(mapping, f, indent=2, ensure_ascii=False)

def get_or_create_global_device_id(ip, port):
    key = f"{ip}:{port}"
    mapping = load_device_map()
    if key in mapping:
        return mapping[key]
    new_id = max(mapping.values(), default=0) + 1
    mapping[key] = new_id
    save_device_map(mapping)
    return new_id

# ---------- 设备（按用户隔离） ----------
def load_devices(username):
    user_dir = get_user_dir(username)
    dev_file = os.path.join(user_dir, "devices.json")
    if not os.path.exists(dev_file):
        save_devices(username, [])
        return []
    with open(dev_file, "r", encoding="utf-8") as f:
        return json.load(f)

def save_devices(username, devs):
    user_dir = get_user_dir(username)
    dev_file = os.path.join(user_dir, "devices.json")
    with open(dev_file, "w", encoding="utf-8") as f:
        json.dump(devs, f, indent=2, ensure_ascii=False)

# ---------- 区域（按用户隔离） ----------
def load_areas(username):
    user_dir = get_user_dir(username)
    area_file = os.path.join(user_dir, "areas.json")
    if not os.path.exists(area_file):
        areas = ["传动轴"]
        save_areas(username, areas)
        return areas
    with open(area_file, "r", encoding="utf-8") as f:
        return json.load(f)

def save_areas(username, areas):
    user_dir = get_user_dir(username)
    area_file = os.path.join(user_dir, "areas.json")
    with open(area_file, "w", encoding="utf-8") as f:
        json.dump(areas, f, indent=2, ensure_ascii=False)

# ---------- 人员（公用） ----------
def load_persons():
    if not os.path.exists(PERSONS_FILE):
        save_persons([])
        return []
    with open(PERSONS_FILE, "r", encoding="utf-8") as f:
        persons = json.load(f)
    # 迁移：status 从单一值转为 dict（按设备存储）
    modified = False
    for p in persons:
        if not isinstance(p.get("status"), dict):
            doors = p.get("doors", [])
            old_status = p.get("status", 0)
            p["status"] = {str(d): old_status for d in doors} if doors else {}
            modified = True
    if modified:
        save_persons(persons)
    return persons

def get_person_status(p, did):
    """读取某人在指定设备上的 status"""
    s = p.get("status", {})
    if isinstance(s, dict):
        return s.get(str(did), 0)
    return int(s) if s is not None else 0

def set_person_status(p, did, value):
    """设置某人在指定设备上的 status"""
    if not isinstance(p.get("status"), dict):
        p["status"] = {}
    p["status"][str(did)] = value

def save_persons(persons):
    with open(PERSONS_FILE, "w", encoding="utf-8") as f:
        json.dump(persons, f, indent=2, ensure_ascii=False)

def get_current_user():
    username = session.get("username")
    if not username:
        raise Exception("未登录")
    return username

# ================= 登录验证 =================
@app.before_request
def auto_login_from_cookie():
    """如果 auth cookie 有效但 session 中无用户名，自动恢复 session"""
    if "username" not in session and request.cookies.get("auth"):
        username = request.cookies.get("auth")
        users = load_users()
        if username in users:
            session["username"] = username
            session.permanent = True
            
@app.before_request
def check_login():
    # 原有的登录检查逻辑不变
    if request.path in ["/", "/api/login", "/api/register", "/api/health"]:
        return
    if request.path.startswith("/api/"):
        if "username" not in session:
            return err("未登录", 401)

# ================= 用户管理 =================
@app.route("/api/register", methods=["POST"])
def register():
    data = request.get_json()
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    if not username or not password:
        return err("用户名和密码不能为空")
    users = load_users()
    if username in users:
        return err("用户名已存在")

    # 角色：仅允许管理员（通过 /api/admin/users 接口）创建 admin；
    # 普通注册一律为 user
    role = "user"

    users[username] = {
        "password": generate_password_hash(password),
        "role": role
    }
    save_users(users)

    try:
        user_dir = get_user_dir(username)
        if not os.path.exists(os.path.join(user_dir, "devices.json")):
            save_devices(username, [])
        if not os.path.exists(os.path.join(user_dir, "areas.json")):
            save_areas(username, ["A区", "B区", "C区"])
    except Exception as e:
        print(f"用户 {username} 数据初始化失败: {e}")

    return ok({"message": "注册成功"})

@app.route("/api/login", methods=["POST"])
def login():
    data = request.get_json()
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    users = load_users()
    if username not in users:
        return err("用户名或密码错误", 401)

    pw_hash = get_user_password_hash(username)
    if not check_password_hash(pw_hash, password):
        return err("用户名或密码错误", 401)

    role = get_user_role(username)

    # 设置服务器 session
    session["username"] = username
    session.permanent = True

    resp = ok({"username": username, "role": role})
    resp.set_cookie(
        "auth",
        username,
        max_age=30 * 24 * 3600,
        httponly=False,
        samesite="Lax",
        path="/"
    )
    return resp

@app.route("/api/logout", methods=["POST"])
def logout():
    session.pop("username", None)
    return ok()

@app.route("/api/user", methods=["GET"])
def current_user():
    username = session.get("username")
    role = get_user_role(username) if username else None
    return ok({"username": username, "role": role})

# ================= 设备 API（全局统一ID，用户隔离） =================

# 设备在线缓存：(ip,port) -> (timestamp, is_online)，5秒过期
_device_online_cache = {}
_device_online_ttl = 5  # 缓存有效期（秒）
_device_online_check_timeout = 1.5  # 单次TCP连接超时（秒）

def _check_device_online_cached(ip, port):
    """带缓存的设备在线检测，5秒内复用上次结果"""
    key = (ip, port)
    now = time.time()
    if key in _device_online_cache:
        ts, status = _device_online_cache[key]
        if now - ts < _device_online_ttl:
            return status
    status = check_device_online(ip, port, timeout=_device_online_check_timeout)
    _device_online_cache[key] = (now, status)
    return status

@app.route("/api/devices", methods=["GET"])
def get_devices():
    username = get_current_user()
    devs = load_devices(username)
    # 并行检测所有设备在线状态（5秒缓存 + 线程池并发）
    with ThreadPoolExecutor(max_workers=min(len(devs), 20)) as ex:
        futures = {ex.submit(_check_device_online_cached, d["ip"], d.get("port", 37777)): d for d in devs}
        for f in as_completed(futures):
            futures[f]["online"] = f.result()
    return ok(devs)

@app.route("/api/devices", methods=["POST"])
def add_device():
    username = get_current_user()
    if not is_admin(username):
        return err("无权限：仅管理员可添加设备", 403)
    data = request.get_json()
    if not data.get("name") or not data.get("ip"):
        return err("name 和 ip 必填")

    ip = data["ip"].strip()
    port = int(data.get("port", 37777))

    # 全局唯一ID
    dev_id = get_or_create_global_device_id(ip, port)

    devs = load_devices(username)

    # 该用户是否已添加此设备
    existing = next((d for d in devs if d["id"] == dev_id), None)
    if existing:
        existing.update({
            "name": data["name"],
            "username": data.get("username", "admin"),
            "password": data.get("password", ""),
            "area": data.get("area", "A区"),
            "note": data.get("note", "")
        })
        save_devices(username, devs)
        existing = next((d for d in devs if d["id"] == dev_id), None)
    if existing:
        existing.update({
            "name": data["name"],
            "username": data.get("username", "admin"),
            "password": data.get("password", ""),
            "area": data.get("area", "A区"),
            "note": data.get("note", "")
        })
        save_devices(username, devs)
        # 同步给其他用户
        sync_device_across_users(dev_id, existing)
        return ok(existing)

    new_dev = {
        "id": dev_id,
        "name": data["name"],
        "ip": ip,
        "port": port,
        "username": data.get("username", "admin"),
        "password": data.get("password", ""),
        "area": data.get("area", "A区"),
        "note": data.get("note", "")
    }
    devs.append(new_dev)
    save_devices(username, devs)
    return ok(new_dev)

@app.route("/api/devices/<int:device_id>", methods=["PUT"])
def update_device(device_id):
    username = get_current_user()
    if not is_admin(username):
        return err("无权限：仅管理员可修改设备", 403)
    data = request.get_json()
    devs = load_devices(username)
    for d in devs:
        if d["id"] == device_id:
            d.update({
                "name": data.get("name", d["name"]),
                "ip": data.get("ip", d["ip"]),
                "port": int(data.get("port", d["port"])),
                "username": data.get("username", d["username"]),
                "password": data.get("password", d["password"]),
                "area": data.get("area", d.get("area", "A区")),
                "note": data.get("note", d.get("note", ""))
            })
            save_devices(username, devs)
            # 同步给其他用户
            sync_device_across_users(device_id, d)
            return ok(d)
    return err("设备不存在", 404)

@app.route("/api/devices/<int:device_id>", methods=["DELETE"])
def delete_device(device_id):
    username = get_current_user()
    if not is_admin(username):
        return err("无权限：仅管理员可删除设备", 403)
    devs = load_devices(username)
    new_devs = [d for d in devs if d["id"] != device_id]
    if len(new_devs) == len(devs):
        return err("设备不存在", 404)
    save_devices(username, new_devs)
    return ok({"deleted_id": device_id})

# ================= 区域 API（用户隔离） =================
@app.route("/api/areas", methods=["GET"])
def get_areas():
    username = get_current_user()
    return ok(load_areas(username))

@app.route("/api/areas", methods=["POST"])
def add_area():
    username = get_current_user()
    name = request.json.get("name", "").strip()
    if not name:
        return err("区域名称不能为空")
    areas = load_areas(username)
    if name in areas:
        return err("区域已存在")
    areas.append(name)
    save_areas(username, areas)
    return ok(areas)

@app.route("/api/areas/<name>", methods=["DELETE"])
def delete_area(name):
    username = get_current_user()
    areas = load_areas(username)
    if name not in areas:
        return err("区域不存在", 404)
    devs = load_devices(username)
    if any(d.get("area") == name for d in devs):
        return err("该区域下有设备，无法删除")
    areas.remove(name)
    save_areas(username, areas)
    return ok({"deleted": name})

# ================= 人员 API（公用） =================
@app.route("/api/persons", methods=["GET"])
def get_persons():
    device_id = request.args.get("device_id")
    persons = load_persons()
    if device_id:
        did = int(device_id)
        persons = [p for p in persons if did in p.get("doors", [])]
    else:
        # 全部设备：只返回当前用户设备中的人员（本地查找，按用户设备过滤）
        username = get_current_user()
        user_devices = load_devices(username)
        user_device_ids = {d["id"] for d in user_devices}
        persons = [p for p in persons if any(did in p.get("doors", []) for did in user_device_ids)]
    return ok(persons)

@app.route("/api/persons/import", methods=["POST"])
def import_persons():
    data = request.get_json()
    device_id = data.get("device_id")          # 全局ID
    new_persons = data.get("persons", [])
    if device_id is None or not isinstance(new_persons, list):
        return err("参数错误")

    persons = load_persons()
    index = {p["user_id"]: p for p in persons}

    for np in new_persons:
        uid = np["user_id"]
        if uid in index:
            existing = index[uid]
            doors = existing.setdefault("doors", [])
            if device_id not in doors:
                doors.append(device_id)
                # 这个门的新人员，默认该设备已有人脸
                existing.setdefault("has_face", {})[str(device_id)] = True
            existing.update({
                "name": np.get("name", existing["name"]),
                "valid_begin": np.get("valid_begin", existing.get("valid_begin")),
                "valid_end": np.get("valid_end", existing.get("valid_end")),
            })
            # status: 兼容旧格式（int）和新格式（dict）
            imported_status = np.get("status")
            if imported_status is not None:
                existing.setdefault("status", {})
                if isinstance(imported_status, dict):
                    existing["status"].update({str(k): int(v) for k, v in imported_status.items()})
                else:
                    existing["status"][str(device_id)] = int(imported_status)
            if np.get("has_face") is not None:
                # 兼容旧格式（bool）和新格式（dict）
                existing.setdefault("has_face", {})
                hf = np["has_face"]
                if isinstance(hf, dict):
                    existing["has_face"].update(hf)
                elif hf:
                    existing["has_face"][str(device_id)] = True
        else:
            np.setdefault("doors", [device_id])
            # 从设备拉取的新人员默认该设备已有人脸
            np["has_face"] = {str(device_id): True}
            # status: 按设备存储
            raw_status = np.pop("status", 0) if "status" in np else 0
            if isinstance(raw_status, dict):
                np["status"] = {str(k): int(v) for k, v in raw_status.items()}
            else:
                np["status"] = {str(device_id): int(raw_status) if raw_status is not None else 0}
            persons.append(np)
            index[uid] = np

    save_persons(persons)
    # 如果设备返回了人员ID列表，清理不在设备上的人的门权限
    imported_uids = {np["user_id"] for np in new_persons}
    changed = False
    for p in persons:
        if p["user_id"] not in imported_uids and device_id in p.get("doors", []):
            p["doors"].remove(device_id)
            p.get("has_face", {}).pop(str(device_id), None)
            p.get("status", {}).pop(str(device_id), None)
            changed = True
    if changed:
        save_persons(persons)
    return ok({"count": len(new_persons)})

@app.route("/api/persons/<user_id>", methods=["DELETE"])
def delete_person(user_id):
    persons = load_persons()
    new_persons = [p for p in persons if p["user_id"] != user_id]
    if len(new_persons) == len(persons):
        return err("人员不存在", 404)
    save_persons(new_persons)
    return ok()

@app.route("/api/persons/<user_id>", methods=["PUT"])
def update_person(user_id):
    data = request.get_json()
    persons = load_persons()
    for p in persons:
        if p["user_id"] == user_id:
            if "has_face" in data:
                p["has_face"] = data["has_face"]
            # 设备级人脸状态更新
            if "has_face_device" in data and "has_face_value" in data:
                p.setdefault("has_face", {})
                p["has_face"][str(data["has_face_device"])] = data["has_face_value"]
            p.update({
                "name": data.get("name", p.get("name")),
                "valid_begin": data.get("valid_begin", p.get("valid_begin")),
                "valid_end": data.get("valid_end", p.get("valid_end")),
                "doors": data.get("doors", p.get("doors")),
            })
            # status: 兼容 dict 和旧格式 int
            if "status" in data:
                imported_status = data["status"]
                p.setdefault("status", {})
                if isinstance(imported_status, dict):
                    p["status"].update({str(k): int(v) for k, v in imported_status.items()})
                else:
                    # 旧格式：整数值应用到所有 doors
                    for d in p.get("doors", []):
                        p["status"][str(d)] = int(imported_status)
            save_persons(persons)
            return ok(p)
    return err("人员不存在", 404)

@app.route("/api/batch_import", methods=["POST"])
def batch_import():
    try:
        username = get_current_user()
    except Exception as e:
        return err(str(e), 401)

    if 'files' not in request.files:
        return err("请选择文件夹上传")

    uploaded_files = request.files.getlist('files')

    user_devices = load_devices(username)
    device_map = {}
    for d in user_devices:
        device_map[d['name']] = d
        device_map[f"{d['ip']}:{d['port']}"] = d

    excel_data = None
    images = {}
    for f in uploaded_files:
        basename = os.path.basename(f.filename)
        if basename == 'user.xlsx':
            excel_data = f.read()
        elif f.filename.lower().endswith(('.jpg','.jpeg','.png','.bmp')):
            images[basename] = f.read()

    if excel_data is None:
        return err("文件夹中未找到 user.xlsx 文件")

    try:
        wb = load_workbook(io.BytesIO(excel_data))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            return err("Excel 数据不足（需要第一行说明、第二行列名、至少一行数据）")
        headers = [str(c).strip() if c else '' for c in rows[1]]
        required = ['用户编号', '姓名']
        for r in required:
            if r not in headers:
                return err(f"Excel 缺少必需列: {r}")
        data_rows = rows[2:]
    except Exception as e:
        return err(f"Excel 解析失败: {e}")

    total = 0
    success = 0
    fail = 0
    face_success = 0
    face_fail = 0
    details = []          # 存储每条记录的结果

    for row in data_rows:
        # 跳过完全为空的行（所有单元格均为 None 或空字符串）
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        # 提取用户编号和姓名，如果无效则跳过该行
        try:
            user_id = str(row[headers.index('用户编号')]).strip()
            name = str(row[headers.index('姓名')]).strip()
        except (ValueError, IndexError):
            continue
        if not user_id or not name:
            continue

        total += 1

        valid_begin = "2000-01-01"
        valid_end = None
        if '有效期结束' in headers:
            ve = row[headers.index('有效期结束')]
            if ve is not None:
                ve_str = str(ve).strip()
                if ve_str and ve_str != 'None':
                    # 尝试多种日期格式，转换为标准 YYYY-MM-DD
                    for fmt in ("%Y-%m-%d", "%Y%m%d", "%Y/%m/%d", "%Y.%m.%d"):
                        try:
                            valid_end = datetime.strptime(ve_str, fmt).strftime("%Y-%m-%d")
                            break
                        except ValueError:
                            continue
        # 如果未能成功解析或未填写，默认 10 年
        if not valid_end:
            valid_end = (datetime.now() + timedelta(days=3650)).strftime("%Y-%m-%d")

        face_filename = ''
        if '人脸图片名称' in headers:
            idx = headers.index('人脸图片名称')
            raw = row[idx]
            face_filename = str(raw).strip() if raw else ''

        door_names = []
        if '门' in headers:
            idx = headers.index('门')
            raw = row[idx]
            if raw:
                door_names = [d.strip() for d in str(raw).split(',') if d.strip()]

        door_ids = []
        resolved_doors = []     # 用于详情显示的门名称列表
        if door_names:
            # 表格中指定了设备名称，只在这些设备上添加
            for dn in door_names:
                dev = device_map.get(dn)
                if not dev:
                    # 尝试模糊匹配
                    for d in user_devices:
                        if dn in d['name'] or dn == f"{d['ip']}:{d['port']}":
                            dev = d
                            break
                if dev:
                    if dev['id'] not in door_ids:  # 避免重复
                        door_ids.append(dev['id'])
                        resolved_doors.append(dev['name'])
                else:
                    # 设备未找到，记录失败信息，但不中断整行
                    resolved_doors.append(f"{dn}(未找到)")
        else:
            # 未指定任何设备，不执行添加，直接计入失败
            fail += 1
            details.append({
                "user_id": user_id,
                "name": name,
                "doors": "未指定设备",
                "status": "失败",
                "face_result": "—",
                "error": "Excel中未填写目标设备"
            })
            continue    # 跳过该行，不再使用当前选中设备

        if not door_ids:
            # 所有指定设备都未找到
            fail += 1
            details.append({
                "user_id": user_id,
                "name": name,
                "doors": ", ".join(resolved_doors) if resolved_doors else "无",
                "status": "失败",
                "face_result": "—",
                "error": "未找到有效设备"
            })
            continue

        # 每行的人员添加可能成功部分门，我们记录整体状态
        row_success = 0
        row_fail = 0
        row_face_ok = 0
        row_face_fail = 0
        error_messages = []

        for did in door_ids:
            dev = next((d for d in user_devices if d['id'] == did), None)
            if not dev:
                row_fail += 1
                error_messages.append(f"设备{did}不存在")
                continue

            try:
                manager.get(dev).add_user(user_id, name)

                persons = load_persons()
                index = {p["user_id"]: p for p in persons}
                new_person = {
                    "user_id": user_id,
                    "name": name,
                    "valid_begin": valid_begin,
                    "valid_end": valid_end,
                    "status": {str(did): 0},
                    "doors": [did],
                    "has_face": {}
                }
                if user_id in index:
                    existing = index[user_id]
                    dlist = existing.setdefault("doors", [])
                    if did not in dlist:
                        dlist.append(did)
                else:
                    persons.append(new_person)
                save_persons(persons)
                row_success += 1

                # 人脸下发
                if face_filename and face_filename in images:
                    # 检查该设备是否已有脸，已有则跳过下发
                    existing_person = next((p for p in persons if p["user_id"] == user_id), None)
                    if existing_person and existing_person.get("has_face", {}).get(str(did)):
                        row_face_ok += 1  # 已有，视为成功
                    else:
                        img_bytes = images[face_filename]
                        tmp_path = os.path.join(BASE, f"_batch_{user_id}.jpg")
                        with open(tmp_path, 'wb') as f_img:
                            f_img.write(img_bytes)
                        try:
                            manager.get(dev).add_face(user_id, tmp_path)
                            # 缓存到本地
                            face_dir = os.path.join(BASE, "faces")
                            os.makedirs(face_dir, exist_ok=True)
                            import shutil
                            shutil.copy(tmp_path, os.path.join(face_dir, f"{user_id}.jpg"))
                            for p in persons:
                                if p["user_id"] == user_id:
                                    p.setdefault("has_face", {})[str(did)] = True
                                    break
                            save_persons(persons)
                            row_face_ok += 1
                        except:
                            row_face_fail += 1
                        finally:
                            if os.path.exists(tmp_path):
                                os.remove(tmp_path)
            except Exception as e:
                row_fail += 1
                error_messages.append(str(e))

        # 汇总该行结果
        status_text = "成功"
        if row_success == 0:
            status_text = "失败"
        elif row_fail > 0:
            status_text = "部分成功"

        face_text = "—"
        if face_filename:
            if row_face_ok == len(door_ids):
                face_text = "全部成功"
            elif row_face_ok > 0:
                face_text = f"部分成功({row_face_ok}/{len(door_ids)})"
            else:
                face_text = "失败"

        if row_success > 0:
            success += row_success
        if row_fail > 0:
            fail += row_fail
        face_success += row_face_ok
        face_fail += row_face_fail

        details.append({
            "user_id": user_id,
            "name": name,
            "doors": ", ".join(resolved_doors) if resolved_doors else "无",
            "status": status_text,
            "face_result": face_text,
            "error": "; ".join(error_messages) if error_messages else ""
        })

    msg = f"处理 {total} 条，成功添加 {success} 人"
    if face_success + face_fail > 0:
        msg += f"，人脸下发 {face_success} 成功 / {face_fail} 失败"
    return ok({
        "total": total,
        "success": success,
        "fail": fail,
        "face_success": face_success,
        "face_fail": face_fail,
        "msg": msg,
        "details": details
    })
    
# 从指定设备移除人员关联（全局ID）
@app.route("/api/persons/<user_id>/devices/<int:device_id>", methods=["DELETE"])
def remove_person_from_device(user_id, device_id):
    persons = load_persons()
    for p in persons:
        if p["user_id"] == user_id:
            doors = p.get("doors", [])
            if device_id in doors:
                doors.remove(device_id)
                if not doors:
                    persons.remove(p)
                else:
                    p["doors"] = doors
                save_persons(persons)
                return ok({"removed_device": device_id, "remaining_doors": doors})
            else:
                return err("该人员未关联此设备", 404)
    return err("人员不存在", 404)

# ================= 设备操作（门禁/日志/人脸） =================
def extract_device():
    data = {}
    if request.is_json:
        data.update(request.get_json(silent=True) or {})
    if request.form:
        data.update(request.form.to_dict())
    if request.args:
        for k, v in request.args.items():
            if k not in data:
                data[k] = v
    ip = data.get("device_ip")
    username = data.get("username")
    password = data.get("password")
    if not ip or not username or not password:
        raise Exception("缺少 device_ip / username / password")
    return {
        "ip": ip,
        "port": int(data.get("device_port", 37777)),
        "username": username,
        "password": password,
    }

@app.route("/api/open", methods=["POST"])
def open_door():
    try:
        dev = extract_device()
        channel = int((request.get_json() or {}).get("channel", 0))
        manager.get(dev).open_door(channel)
        return ok({"channel": channel})
    except Exception as e:
        return err(e, 500)

@app.route("/api/door/status", methods=["GET"])
def door_status():
    try:
        dev = extract_device()
        channel = int(request.args.get("channel", 1))
        status = manager.get(dev).get_door_status(channel)
        if status is None:
            return err("查询门状态失败", 500)
        is_open = status.lower() == "open"
        return ok({"status": status, "is_open": is_open})
    except Exception as e:
        return err(str(e), 500)

@app.route("/api/user", methods=["POST"])
def add_user():
    try:
        dev = extract_device()
        data = request.get_json()
        manager.get(dev).add_user(data["user_id"], data["name"])
        return ok()
    except Exception as e:
        return err(e, 500)

@app.route("/api/user/<uid>", methods=["PUT"])
def update_user(uid):
    try:
        dev = extract_device()
        data = request.get_json(silent=True) or {}
        name = (data.get("name") or "").strip()
        if not name:
            return err("缺少 name", 400)

        doors = data.get("doors")
        if doors is not None:
            if not isinstance(doors, list) or not doors:
                return err("doors 必须是非空数组", 400)
            doors = [int(item) for item in doors]

        valid_begin = (data.get("valid_begin") or "").strip() or None
        valid_end = (data.get("valid_end") or "").strip() or None
        status = data.get("status")
        if status is not None:
            status = int(status)

        manager.get(dev).update_user(
            uid,
            name,
            status=0 if status is None else status,
            doors=doors,
            valid_begin=valid_begin,
            valid_end=valid_end,
        )
        return ok({
            "user_id": uid,
            "name": name,
            "doors": doors,
            "valid_begin": valid_begin,
            "valid_end": valid_end,
            "status": 0 if status is None else status,
        })
    except Exception as e:
        return err(e, 500)

@app.route("/api/user/<uid>", methods=["DELETE"])
def del_user(uid):
    try:
        dev = extract_device()
        manager.get(dev).delete_user(uid)
        return ok()
    except Exception as e:
        return err(e, 500)

@app.route("/api/user/<uid>/freeze", methods=["POST"])
def freeze_user(uid):
    try:
        dev = extract_device()
        manager.get(dev).freeze_user(uid)
        return ok()
    except Exception as e:
        return err(e, 500)

@app.route("/api/user/<uid>/unfreeze", methods=["POST"])
def unfreeze_user(uid):
    try:
        dev = extract_device()
        manager.get(dev).unfreeze_user(uid)
        return ok()
    except Exception as e:
        return err(e, 500)

@app.route("/api/user/<uid>/validity", methods=["PUT"])
def update_user_validity(uid):
    try:
        dev = extract_device()
        data = request.get_json(silent=True) or {}
        valid_begin = (data.get("valid_begin") or "").strip()
        valid_end = (data.get("valid_end") or "").strip()
        if not valid_begin or not valid_end:
            return err("缺少 valid_begin / valid_end", 400)
        manager.get(dev).update_user_validity(uid, valid_begin, valid_end)
        return ok({"valid_begin": valid_begin, "valid_end": valid_end})
    except Exception as e:
        return err(e, 500)

@app.route("/api/device/user/id/<uid>", methods=["GET"])
def get_device_user_by_id(uid):
    try:
        dev = extract_device()
        user = manager.get(dev).get_user_by_id(uid)
        if user:
            return ok(user)
        return err("用户不存在", 404)
    except Exception as e:
        return err(str(e), 500)

@app.route("/api/device/users/search", methods=["GET"])
def search_device_users():
    try:
        dev = extract_device()
        keyword = request.args.get("keyword", "").strip()
        if not keyword:
            return err("缺少 keyword")
        users = manager.get(dev).search_users_by_name(keyword)
        return ok({"total": len(users), "items": users})
    except Exception as e:
        return err(str(e), 500)

@app.route("/api/device/users/all", methods=["GET"])
def get_all_device_users():
    try:
        dev = extract_device()
        page = int(request.args.get("page", 1))
        page_size = int(request.args.get("page_size", 20))
        if page < 1: page = 1
        if page_size <= 0: page_size = 10000  # 获取全部
        offset = (page - 1) * page_size
        total, users = manager.get(dev).get_users_paginated(offset, page_size)
        return ok({"total": total, "page": page, "page_size": page_size, "items": users})
    except Exception as e:
        return err(str(e), 500)

@app.route("/api/face/<uid>", methods=["POST"])
def add_face(uid):
    try:
        dev = extract_device()
        did = get_or_create_global_device_id(dev["ip"], dev["port"])

        # 检查本地状态：该设备已有脸则跳过下发（force=1 时强制下发）
        force = request.form.get("force")
        if force != "1":
            persons = load_persons()
            for p in persons:
                if p["user_id"] == uid:
                    if p.get("has_face", {}).get(str(did)):
                        print(f"[FACE] SKIP uid={uid} device={did} — 本地记录已有脸")
                        return ok()  # 已有，跳过
                    break

        print(f"[FACE] SEND uid={uid} device={did} force={force} — 开始下发")

        f = request.files["file"]
        path = os.path.join(BASE, f"_tmp_{uid}.jpg")
        f.save(path)
        manager.get(dev).add_face(uid, path)
        # 缓存到本地，方便后续同步到其他设备
        face_dir = os.path.join(BASE, "faces")
        os.makedirs(face_dir, exist_ok=True)
        import shutil
        shutil.copy(path, os.path.join(face_dir, f"{uid}.jpg"))
        # 更新 persons.json 中该设备的 has_face 状态
        # 重新加载（上面的 persons 可能已过期）
        persons = load_persons()
        for p in persons:
            if p["user_id"] == uid:
                p.setdefault("has_face", {})[str(did)] = True
                break
        save_persons(persons)
        if os.path.exists(path):
            os.remove(path)
        return ok()
    except Exception as e:
        return err(e, 500)

@app.route("/api/face/<uid>", methods=["GET"])
def get_cached_face(uid):
    """获取本地缓存的人脸照片"""
    try:
        cache_path = os.path.join(BASE, "faces", f"{uid}.jpg")
        if not os.path.exists(cache_path):
            return err("未找到人脸缓存", 404)
        with open(cache_path, "rb") as f:
            return Response(f.read(), mimetype="image/jpeg")
    except Exception as e:
        return err(e, 500)

@app.route("/api/face/<uid>/exists", methods=["GET"])
def face_exists(uid):
    """检查本地 faces 文件夹是否有此人照片"""
    for ext in (".jpg", ".png", ".jpeg"):
        if os.path.exists(os.path.join(BASE, "faces", f"{uid}{ext}")):
            return jsonify({"exists": True})
    return jsonify({"exists": False})

@app.route("/api/face/<uid>", methods=["DELETE"])
def del_face(uid):
    try:
        dev = extract_device()
        manager.get(dev).delete_face(uid)
        # 更新 persons.json 中该设备的 has_face 状态
        persons = load_persons()
        did = get_or_create_global_device_id(dev["ip"], dev["port"])
        for p in persons:
            if p["user_id"] == uid:
                p.setdefault("has_face", {})[str(did)] = False
                break
        save_persons(persons)
        return ok()
    except Exception as e:
        return err(e, 500)

@app.route("/api/log", methods=["GET"])
def log():
    try:
        dev = extract_device()
        start_str = request.args.get("start")
        end_str = request.args.get("end")
        if not start_str or not end_str:
            return err("缺少 start 或 end 参数")
        try:
            start = datetime.strptime(start_str, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            start = datetime.strptime(start_str, "%Y-%m-%d")
        try:
            end = datetime.strptime(end_str, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            end = datetime.strptime(end_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        records = manager.get(dev).query_log(start, end)
        return ok({"total": len(records), "records": records})
    except Exception as e:
        return err(e, 500)

@app.route("/api/health", methods=["GET"])
def health():
    return ok({"status": "running", "time": datetime.now().isoformat()})

# ================= WebSocket 视频预览 =================
# 全局预览会话管理：device_id -> {ws_set, client, ffmpeg_proc, thread}
_preview_sessions = {}
_preview_lock = threading.Lock()


def _preview_key(device_id):
    return str(device_id)


@sock.route("/ws/preview/<int:device_id>")
def ws_preview(ws, device_id):
    """
    WebSocket 视频预览端点。
    每个设备只启动一个 SDK 预览 + 一个 ffmpeg 进程，
    多个客户端连接同一设备时共享同一路码流。
    """
    # 从 cookie 验证登录（WebSocket 握手携带 cookie）
    username = session.get("username")
    if not username:
        ws.send(json.dumps({"type": "error", "msg": "未登录"}))
        return

    # 查找设备信息（遍历所有用户，只要 device_id 匹配即可）
    device_info = _find_device_by_id(device_id, username)
    if not device_info:
        ws.send(json.dumps({"type": "error", "msg": "设备不存在或无权限"}))
        return

    key = _preview_key(device_id)

    with _preview_lock:
        if key not in _preview_sessions:
            # 第一个连接：创建预览会话
            try:
                session_obj = _create_preview_session(device_id, device_info)
                _preview_sessions[key] = session_obj
                print(f"[ws_preview] 新建预览会话 device_id={device_id}")
            except Exception as e:
                ws.send(json.dumps({"type": "error", "msg": f"启动预览失败: {e}"}))
                return
        else:
            session_obj = _preview_sessions[key]

        session_obj["ws_set"].add(ws)
        print(f"[ws_preview] 客户端加入 device_id={device_id} 当前连接数={len(session_obj['ws_set'])}")

    # 新客户端：先发送 jsmp header + 预缓冲数据（JSMpeg 要求首条消息包含头信息）
    header = session_obj.get("jsmp_header", b"")
    pre_buf = session_obj.get("pre_buffer", b"")
    first_msg = header + pre_buf
    if first_msg:
        try:
            ws.send(first_msg)
        except Exception:
            print(f"[ws_preview] 发送初始数据失败 device_id={device_id}")
            # 发送失败也算断开，从 ws_set 移除
            with _preview_lock:
                if key in _preview_sessions:
                    _preview_sessions[key]["ws_set"].discard(ws)
            return

    try:
        # 保持连接，等待客户端断开
        while True:
            try:
                msg = ws.receive(timeout=30)
                if msg is None:
                    break
                # 收到 ping 就 pong，保持心跳
                if msg == "ping":
                    ws.send("pong")
            except Exception:
                break
    finally:
        with _preview_lock:
            if key in _preview_sessions:
                session_obj = _preview_sessions[key]
                session_obj["ws_set"].discard(ws)
                remaining = len(session_obj["ws_set"])
                print(f"[ws_preview] 客户端离开 device_id={device_id} 剩余连接数={remaining}")
                if remaining == 0:
                    _stop_preview_session(key, session_obj)
                    del _preview_sessions[key]
                    print(f"[ws_preview] 预览会话已关闭 device_id={device_id}")


def _find_device_by_id(device_id, username):
    """在用户的设备列表中查找 device_id 对应的设备信息"""
    devs = load_devices(username)
    return next((d for d in devs if d["id"] == device_id), None)


def _parse_mpeg_sequence_header(data):
    """在 MPEG-1 视频 ES 数据中查找序列头 (00 00 01 B3)，返回 (width, height, offset) 或 (None, None, -1)"""
    idx = data.find(b'\x00\x00\x01\xb3')
    if idx < 0 or idx + 8 > len(data):
        return None, None, -1
    # MPEG-1 sequence header layout (after 4-byte start code):
    # 12 bits: horizontal_size
    # 12 bits: vertical_size
    hdr = data[idx + 4:idx + 8]
    width = (hdr[0] << 4) | (hdr[1] >> 4)
    height = ((hdr[1] & 0x0F) << 8) | hdr[2]
    return width, height, idx


def _create_preview_session(device_id, device_info):
    """
    创建预览会话：
    1. 启动 ffmpeg 子进程（stdin=pipe 读 dhav，stdout=pipe 输出 mpeg1video ES）
    2. 启动 SDK 预览，把数据写入 ffmpeg stdin
    3. 从 ffmpeg stdout 解析 MPEG 序列头获取分辨率，构建 jsmp header
    4. 启动推流线程，把 ffmpeg stdout 广播给所有 WebSocket 客户端
    """
    ffmpeg_cmd = [
        "ffmpeg", "-loglevel", "error",
        "-f", "dhav",           # 输入格式：大华私有容器
        "-i", "pipe:0",         # 从 stdin 读
        "-f", "mpeg1video",     # 输出格式：原始 MPEG-1 视频 ES（JSMpeg 要求）
        "-vcodec", "mpeg1video",
        "-b:v", "600k",
        "-r", "25",
        "-bf", "0",             # 禁用 B 帧，降低延迟
        "-an",                  # 无音频（JSMpeg ES 模式不支持音频）
        "pipe:1"                # 输出到 stdout
    ]

    ffmpeg_proc = subprocess.Popen(
        ffmpeg_cmd,
        stdin=subprocess.PIPE,
        stdout=subprocess.PIPE,
        stderr=subprocess.DEVNULL,
    )

    ws_set = set()
    stop_event = threading.Event()

    session_obj = {
        "ws_set": ws_set,
        "ffmpeg_proc": ffmpeg_proc,
        "stop_event": stop_event,
        "device_info": device_info,
        "device_id": device_id,
    }

    # 回调：SDK 数据 → ffmpeg stdin
    def on_sdk_data(data: bytes):
        if stop_event.is_set():
            return
        try:
            ffmpeg_proc.stdin.write(data)
            ffmpeg_proc.stdin.flush()
        except Exception as e:
            print(f"[preview] ffmpeg stdin write error: {e}")
            stop_event.set()

    # 启动 SDK 预览
    client = manager.get(device_info)
    client.start_preview(on_sdk_data, channel=0)
    session_obj["client"] = client

    # 从 ffmpeg stdout 解析 MPEG 序列头，获取分辨率，构建 jsmp header
    import struct as _struct
    initial_buffer = b""
    jsmp_header = None
    pre_buffer = b""
    deadline = time.time() + 10  # 最多等 10 秒

    while time.time() < deadline:
        try:
            chunk = ffmpeg_proc.stdout.read(4096)
        except Exception:
            break
        if not chunk:
            break
        initial_buffer += chunk
        width, height, idx = _parse_mpeg_sequence_header(initial_buffer)
        if width is not None:
            jsmp_header = b"jsmp" + _struct.pack(">HH", width, height)
            pre_buffer = initial_buffer[idx:]  # 从序列头开始
            print(f"[preview] 解析到视频分辨率: {width}x{height} device_id={device_id}")
            break

    if jsmp_header is None:
        # 超时或失败：使用默认分辨率
        print(f"[preview] 未能解析分辨率，使用默认 1280x720 device_id={device_id}")
        jsmp_header = b"jsmp" + _struct.pack(">HH", 1280, 720)
        pre_buffer = initial_buffer if initial_buffer else b""

    session_obj["jsmp_header"] = jsmp_header
    session_obj["pre_buffer"] = pre_buffer

    # 推流线程：ffmpeg stdout → 广播给所有 WebSocket 客户端
    def _broadcast_thread():
        CHUNK = 4096
        while not stop_event.is_set():
            try:
                chunk = ffmpeg_proc.stdout.read(CHUNK)
                if not chunk:
                    print(f"[preview] ffmpeg stdout EOF device_id={device_id}")
                    break
                # 广播给所有客户端（复制一份 ws_set 避免迭代中修改）
                dead = set()
                with _preview_lock:
                    targets = set(ws_set)
                for ws_client in targets:
                    try:
                        ws_client.send(chunk)
                    except Exception:
                        dead.add(ws_client)
                if dead:
                    with _preview_lock:
                        ws_set.difference_update(dead)
            except Exception as e:
                if not stop_event.is_set():
                    print(f"[preview] broadcast error: {e}")
                break
        print(f"[preview] 推流线程退出 device_id={device_id}")

    t = threading.Thread(target=_broadcast_thread, daemon=True, name=f"preview-{device_id}")
    t.start()
    session_obj["broadcast_thread"] = t

    return session_obj


def _stop_preview_session(key, session_obj):
    """停止预览会话：关闭 SDK 预览、终止 ffmpeg"""
    session_obj["stop_event"].set()
    # 停止 SDK 预览
    try:
        client = session_obj.get("client")
        if client:
            client.stop_preview()
    except Exception as e:
        print(f"[preview] stop_preview error: {e}")
    # 关闭 ffmpeg
    try:
        proc = session_obj.get("ffmpeg_proc")
        if proc:
            try:
                proc.stdin.close()
            except Exception:
                pass
            proc.terminate()
            proc.wait(timeout=5)
    except Exception as e:
        print(f"[preview] ffmpeg terminate error: {e}")

def sync_device_across_users(device_id, updated_data):
    """将设备信息同步到所有拥有该设备的用户"""
    if not os.path.exists(DATA_ROOT):
        return
    # 遍历所有用户目录
    for username in os.listdir(DATA_ROOT):
        # 跳过非目录文件
        user_dir = os.path.join(DATA_ROOT, username)
        if not os.path.isdir(user_dir):
            continue
        # 跳过管理员自己（管理员的信息已经由调用方更新）
        try:
            current_user = get_current_user()
            if username == current_user:
                continue
        except:
            pass

        devs = load_devices(username)
        modified = False
        for d in devs:
            if d["id"] == device_id:
                d.update({
                    "name": updated_data.get("name", d["name"]),
                    "ip": updated_data.get("ip", d["ip"]),
                    "port": int(updated_data.get("port", d["port"])),
                    "username": updated_data.get("username", d["username"]),
                    "password": updated_data.get("password", d["password"]),
                    "area": updated_data.get("area", d.get("area", "A区")),
                    "note": updated_data.get("note", d.get("note", ""))
                })
                modified = True
        if modified:
            save_devices(username, devs)

def check_device_online(ip, port, timeout=2):
    try:
        sock = socket.create_connection((ip, port), timeout=timeout)
        sock.close()
        return True
    except Exception:
        return False

@app.route("/download/template")
def download_template():
    template_path = os.path.join(BASE, "user.xlsx")
    if not os.path.exists(template_path):
        return err("模板文件不存在", 404)
    return send_from_directory(
        BASE, "user.xlsx",
        as_attachment=True,
        download_name="user.xlsx"
    )
    
# ================= 管理员：用户管理 API =================

@app.route("/api/admin/users", methods=["GET"])
def admin_list_users():
    try:
        require_admin()
    except Exception as e:
        return err(str(e), 403)
    users = load_users()
    result = []
    for uname, entry in users.items():
        if isinstance(entry, dict):
            role = entry.get("role", "user")
        else:
            role = "user"
        result.append({"username": uname, "role": role})
    return ok(result)

@app.route("/api/admin/users", methods=["POST"])
def admin_create_user():
    try:
        require_admin()
    except Exception as e:
        return err(str(e), 403)
    data = request.get_json()
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    role = data.get("role", "user")
    if not username or not password:
        return err("用户名和密码不能为空")
    if role not in ("admin", "user"):
        return err("角色必须是 admin 或 user")
    users = load_users()
    if username in users:
        return err("用户名已存在")
    users[username] = {"password": generate_password_hash(password), "role": role}
    save_users(users)
    # 初始化用户数据目录
    try:
        user_dir = get_user_dir(username)
        if not os.path.exists(os.path.join(user_dir, "devices.json")):
            save_devices(username, [])
        if not os.path.exists(os.path.join(user_dir, "areas.json")):
            save_areas(username, ["A区", "B区", "C区"])
    except Exception as e:
        print(f"用户 {username} 数据初始化失败: {e}")
    return ok({"username": username, "role": role})

@app.route("/api/admin/users/<target_username>", methods=["PUT"])
def admin_update_user(target_username):
    try:
        require_admin()
    except Exception as e:
        return err(str(e), 403)
    data = request.get_json()
    users = load_users()
    if target_username not in users:
        return err("用户不存在", 404)
    entry = users[target_username]
    if not isinstance(entry, dict):
        entry = {"password": entry, "role": "user"}

    if "role" in data:
        if data["role"] not in ("admin", "user"):
            return err("角色必须是 admin 或 user")
        entry["role"] = data["role"]
    if "password" in data and data["password"].strip():
        entry["password"] = generate_password_hash(data["password"].strip())
    users[target_username] = entry
    save_users(users)
    return ok({"username": target_username, "role": entry["role"]})

@app.route("/api/admin/users/<target_username>", methods=["DELETE"])
def admin_delete_user(target_username):
    try:
        require_admin()
    except Exception as e:
        return err(str(e), 403)
    current = session.get("username")
    if target_username == current:
        return err("不能删除自己")
    users = load_users()
    if target_username not in users:
        return err("用户不存在", 404)
    del users[target_username]
    save_users(users)
    return ok()

@app.route("/api/admin/users/<target_username>/devices", methods=["POST"])
def admin_assign_devices(target_username):
    """管理员将指定设备分配给某个用户（复制管理员的设备配置）"""
    try:
        admin_username = session.get("username")
        require_admin()
    except Exception as e:
        return err(str(e), 403)
    users = load_users()
    if target_username not in users:
        return err("目标用户不存在", 404)
    data = request.get_json()
    device_ids = data.get("device_ids", [])
    if not isinstance(device_ids, list):
        return err("device_ids 必须是数组")

    # 从管理员设备列表中找到对应设备，复制给目标用户
    admin_devs = load_devices(admin_username)
    target_devs = load_devices(target_username)
    target_index = {d["id"]: d for d in target_devs}

    assigned = []
    for did in device_ids:
        src = next((d for d in admin_devs if d["id"] == did), None)
        if not src:
            continue
        if did not in target_index:
            target_devs.append(dict(src))
            assigned.append(did)
        else:
            # 设备已存在，同步管理员修改的字段（备注、名称、IP等）
            target_index[did].update({k: src[k] for k in ('name', 'ip', 'port', 'username', 'password', 'area', 'note') if k in src})
            assigned.append(did)

    save_devices(target_username, target_devs)
    return ok({"assigned": assigned})

@app.route("/api/admin/users/<target_username>/devices", methods=["GET"])
def admin_get_user_devices(target_username):
    """获取某个用户拥有的设备ID列表"""
    try:
        require_admin()
    except Exception as e:
        return err(str(e), 403)
    devs = load_devices(target_username)
    return ok([d["id"] for d in devs])

@app.route("/api/admin/users/<target_username>/devices/<int:device_id>", methods=["DELETE"])
def admin_remove_user_device(target_username, device_id):
    """管理员从目标用户的设备列表中移除某台设备"""
    try:
        require_admin()
    except Exception as e:
        return err(str(e), 403)
    users = load_users()
    if target_username not in users:
        return err("目标用户不存在", 404)
    devs = load_devices(target_username)
    new_devs = [d for d in devs if d["id"] != device_id]
    save_devices(target_username, new_devs)
    return ok({"removed": device_id})

# ================= 静态文件 =================
@app.route("/")
@app.route("/<path:filename>")
def static_files(filename="access_control.html"):
    return send_from_directory(BASE, filename)

# ================= 清理线程 =================
def cleaner():
    while True:
        time.sleep(60)
        manager.cleanup()

threading.Thread(target=cleaner, daemon=True).start()

# ================= 启动 =================
if __name__ == "__main__":
    if not os.path.exists(DATA_ROOT):
        os.makedirs(DATA_ROOT, exist_ok=True)

    # 迁移旧格式 users.json（纯密码哈希字符串 → dict with role）
    users = load_users()
    migrated = False
    first_username = None
    for uname, entry in users.items():
        if first_username is None:
            first_username = uname
        if isinstance(entry, str):
            # 旧格式：首个用户升为 admin，其余为 user
            role = "admin" if uname == first_username else "user"
            users[uname] = {"password": entry, "role": role}
            migrated = True
    if migrated:
        # 确保至少有一个 admin
        has_admin = any(
            (v.get("role") == "admin" if isinstance(v, dict) else False)
            for v in users.values()
        )
        if not has_admin and first_username:
            users[first_username]["role"] = "admin"
        save_users(users)
        print("✅ users.json 已迁移至新格式（含角色字段）")

    parser = argparse.ArgumentParser()
    parser.add_argument("--port", type=int, default=15001)
    args = parser.parse_args()
    app.run("0.0.0.0", args.port, threaded=True)